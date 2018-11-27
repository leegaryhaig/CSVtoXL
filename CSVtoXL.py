import openpyxl, csv, os
from openpyxl.utils import get_column_letter

# Creates wb object
wb = openpyxl.Workbook()

def convertCSV(name):
    # Loop through each row and column of the csvData array in the corresponding excel sheet
    # Excel sheet titles have a max of 31 Characters
    if len(name) >= 31:
        name = name[:31]
    sheetname = wb.create_sheet(name)
    for row_num in range(len(csvData)):
        for col_num in range(len(csvData[row_num]) - 1):
            # print('sheetname.cell(' + str(get_column_letter(col_num + 1)) + str(row_num + 1) + ')' )
            # print('sheetname.cell(' + str(get_column_letter(col_num + 1)) + str(row_num + 1) + ')' + csvData[row_num][col_num])
            sheetname.cell(row=row_num + 1, column=col_num + 1).value = csvData[row_num][col_num]

curDir = []
for files in os.listdir('.'):
    curDir.append(files)

if 'csvfiles' in curDir:
    for root, dirs, files in os.walk('csvfiles'):
        for name in files:
            if name.endswith('.csv'):
                abs_path = os.path.join(root, name)
                sheetname = name.rstrip('.csv')
                csvFile = open(abs_path)
                csvReader = csv.reader(csvFile)
                csvData = list(csvReader)
                convertCSV(name)

    xl_filename = input('Save Excel File as:\n>')
    csvFile.close()
    wb.remove(wb['Sheet'])
    wb.save(str(xl_filename) + '.xlsx')
    wb.close()

else:
    print('can\'t find the csvfiles folder')